# excel_reader.py
import io
<<<<<<< HEAD
import pandas as pd
=======
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
>>>>>>> f116724 (init: estructura Safetti ETL)

class ExcelReadError(RuntimeError):
    pass

<<<<<<< HEAD
=======
def iter_excel_chunks(xls_bytes: bytes, sheet_name=0, skiprows=0, chunk_rows=50000):
    """
    Lee un .xlsx en modo streaming y rinde DataFrames de tamaño <= chunk_rows.
    Asume que la primera fila útil tras skiprows es el header.
    """
    wb = load_workbook(filename=BytesIO(xls_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[sheet_name] if isinstance(sheet_name, int) else wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    # descarta filas previas
    for _ in range(skiprows):
        next(rows_iter, None)

    # header
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return  # hoja vacía

    buffer = []
    for row in rows_iter:
        buffer.append(row)
        if len(buffer) >= chunk_rows:
            yield pd.DataFrame(buffer, columns=header)
            buffer.clear()

    if buffer:
        yield pd.DataFrame(buffer, columns=header)

    wb.close()

>>>>>>> f116724 (init: estructura Safetti ETL)
def read_excel_bytes(
    xls_bytes: bytes,
    *,
    sheet_name=0,      # 0 = primera hoja, o nombre exacto de la hoja
    skiprows=0,        # filas a saltar antes del encabezado
    header=0,          # fila donde están los encabezados (0-based, relativo tras skiprows)
    dtype=None,        # dict opcional de tipos por columna (si quieres forzar)
    usecols=None       # opcional: lista/rango de columnas a leer (p.ej. "A:Z" o ["SKU","Cantidad"])
) -> pd.DataFrame:
    """
    Lee un archivo XLSX desde bytes y devuelve un DataFrame.
    - No modifica nombres de columnas ni tipos, a menos que pases 'dtype'.
    - Si 'usecols' se especifica, solo se cargan esas columnas (más rápido/menos memoria).

    Posibles errores comunes:
      * Hoja no existe -> ExcelReadError con mensaje claro
      * Archivo vacío o corrupto -> ExcelReadError
    """
    try:
        with io.BytesIO(xls_bytes) as bio:
            df = pd.read_excel(
                bio,
                sheet_name=sheet_name,
                skiprows=skiprows,
                header=header,
                dtype=dtype,
                usecols=usecols,
                engine="openpyxl",  # asegúrate de tener openpyxl instalado
            )

        # pd.read_excel devuelve:
        # - DataFrame si sheet_name es int/str
        # - dict[str, DataFrame] si pasas una lista de hojas
        if isinstance(df, dict):
            raise ExcelReadError(
                "Se recibió un dict de DataFrames. Pasa 'sheet_name' como int/str, no como lista."
            )

        # Nota: NO normalizamos columnas (requisito del proyecto)
        return df

    except ValueError as e:
        # ValueError típico cuando la hoja no existe o rango inválido
        raise ExcelReadError(f"Error leyendo Excel: {e}") from e
    except Exception as e:
        raise ExcelReadError(f"Fallo al leer XLSX: {e}") from e
